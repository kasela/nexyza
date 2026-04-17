"""
Advanced export: Excel (with styled sheets), PDF (with charts), PowerPoint.
"""
import io
import json
import math
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.http import HttpResponse
from apps.analyser.models import FileUpload
from .export_context import build_export_context
from apps.whitelabel.branding_engine import build_branding_payload


# ── Chart rendering helper (matplotlib → PNG bytes) ──────────────────────────

def _chart_to_png(chart_cfg, width_in=6.5, height_in=2.8, branding=None) -> bytes | None:
    """Render a ChartConfig's cached data to a PNG using matplotlib."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        import numpy as np

        d = chart_cfg.cached_data or {}
        if not d or d.get('error'):
            return None

        # KPI card: simple text image
        if d.get('kpi') or chart_cfg.chart_type == 'kpi':
            fig, ax = plt.subplots(figsize=(width_in, height_in))
            fig.patch.set_facecolor('#0f0c1e')
            ax.set_facecolor('#0f0c1e')
            ax.axis('off')
            ax.text(0.5, 0.6, str(d.get('value', '—')),
                    ha='center', va='center', fontsize=36, fontweight='bold',
                    color='#a78bfa', transform=ax.transAxes)
            ax.text(0.5, 0.25, str(d.get('label', chart_cfg.title)),
                    ha='center', va='center', fontsize=11, color='#64748b',
                    transform=ax.transAxes)
            buf = io.BytesIO()
            fig.savefig(buf, format='png', bbox_inches='tight', dpi=120,
                        facecolor=fig.get_facecolor())
            plt.close(fig)
            buf.seek(0)
            return buf.getvalue()

        labels   = d.get('labels', [])
        datasets = d.get('datasets', [])
        if not labels or not datasets:
            return None

        chart_type = d.get('chart_type', chart_cfg.chart_type) or 'bar'
        is_horiz   = d.get('is_horizontal', False) or chart_type == 'horizontal_bar'

        # Colour palette — use branding colours first
        p = branding.primary_color if branding else '#7c3aed'
        a = branding.accent_color  if branding else '#3b82f6'
        PALETTE = [p, a, '#10b981', '#f59e0b', '#f43f5e', '#06b6d4',
                   '#ec4899', '#8b5cf6', '#84cc16', '#f97316', '#14b8a6']

        fig, ax = plt.subplots(figsize=(width_in, height_in))
        fig.patch.set_facecolor('#0f0c1e')
        ax.set_facecolor('#111827')
        for spine in ax.spines.values():
            spine.set_edgecolor('#1f2937')
        ax.tick_params(colors='#6b7280', labelsize=8)
        ax.xaxis.label.set_color('#6b7280')
        ax.yaxis.label.set_color('#6b7280')
        ax.grid(color='#1f2937', linewidth=0.5, alpha=0.7)

        x_positions = np.arange(len(labels))

        if chart_type in ('pie', 'doughnut'):
            ax.set_facecolor('#0f0c1e')
            vals = [abs(float(v)) if v is not None else 0
                    for v in (datasets[0].get('data') or [])]
            if not any(vals):
                plt.close(fig); return None
            wedge_kwargs = {'wedgeprops': dict(width=0.5)} if chart_type == 'doughnut' else {}
            wedges, texts, autotexts = ax.pie(
                vals, labels=labels[:len(vals)],
                colors=PALETTE[:len(vals)],
                autopct='%1.1f%%', startangle=90,
                textprops={'color': '#9ca3af', 'fontsize': 8},
                **wedge_kwargs
            )
            for at in autotexts:
                at.set_color('#e5e7eb')
                at.set_fontsize(7)

        elif chart_type == 'scatter':
            for i, ds in enumerate(datasets[:4]):
                pts = ds.get('data') or []
                if pts and isinstance(pts[0], dict):
                    xs = [p.get('x', 0) for p in pts]
                    ys = [p.get('y', 0) for p in pts]
                else:
                    xs, ys = range(len(pts)), [float(v) if v is not None else 0 for v in pts]
                ax.scatter(xs, ys, color=PALETTE[i % len(PALETTE)],
                           s=25, alpha=0.8, label=ds.get('label',''))

        elif chart_type in ('line', 'area'):
            bar_count = len(datasets)
            for i, ds in enumerate(datasets[:6]):
                vals = [float(v) if v is not None else 0 for v in (ds.get('data') or [])]
                color = PALETTE[i % len(PALETTE)]
                ax.plot(x_positions, vals, color=color, linewidth=2,
                        marker='o', markersize=3, label=ds.get('label',''))
                if chart_type == 'area':
                    ax.fill_between(x_positions, vals, alpha=0.15, color=color)

        else:  # bar / horizontal_bar
            bar_count = len(datasets)
            bar_w = 0.8 / max(bar_count, 1)
            offsets = np.linspace(-(bar_w * (bar_count-1)/2),
                                   bar_w * (bar_count-1)/2, bar_count)
            for i, ds in enumerate(datasets[:6]):
                vals = [float(v) if v is not None else 0 for v in (ds.get('data') or [])]
                color = PALETTE[i % len(PALETTE)]
                if is_horiz:
                    ax.barh(x_positions + offsets[i], vals, bar_w * 0.9,
                            color=color, alpha=0.85, label=ds.get('label',''))
                else:
                    ax.bar(x_positions + offsets[i], vals, bar_w * 0.9,
                           color=color, alpha=0.85, label=ds.get('label',''))

        # Axis ticks
        if chart_type not in ('pie', 'doughnut', 'scatter'):
            short_labels = [str(l)[:12] for l in labels]
            if is_horiz:
                ax.set_yticks(x_positions)
                ax.set_yticklabels(short_labels, fontsize=7)
            else:
                ax.set_xticks(x_positions)
                ax.set_xticklabels(short_labels, rotation=30, ha='right',
                                   fontsize=7)

        # Legend for multi-dataset
        if len(datasets) > 1 and chart_type not in ('pie', 'doughnut'):
            ax.legend(fontsize=7, labelcolor='#9ca3af',
                      facecolor='#1f2937', edgecolor='#374151',
                      framealpha=0.8, loc='upper right')

        # Axis labels
        x_lbl = d.get('x_label') or chart_cfg.x_axis
        y_lbl = d.get('y_label') or chart_cfg.y_axis
        if x_lbl and not is_horiz:
            ax.set_xlabel(x_lbl[:30], fontsize=8, color='#6b7280')
        if y_lbl and not is_horiz:
            ax.set_ylabel(y_lbl[:30], fontsize=8, color='#6b7280')

        plt.tight_layout(pad=0.4)
        buf = io.BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight', dpi=130,
                    facecolor=fig.get_facecolor())
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()

    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"Chart render failed: {e}")
        return None


# ── Excel export ──────────────────────────────────────────────────────────────

@login_required
def export_excel(request, pk):
    upload   = get_object_or_404(FileUpload, pk=pk, user=request.user)
    analysis = upload.analysis_result or {}

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl required for Excel export.', status=500)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    hdr_fill = PatternFill('solid', fgColor='4C1D95')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    alt_fill = PatternFill('solid', fgColor='1E1B2E')

    ws['A1'] = upload.original_name
    ws['A1'].font = Font(bold=True, size=14, color='A78BFA')
    ws['A2'] = (f"Rows: {upload.row_count:,}  |  Cols: {upload.column_count}"
                f"  |  Type: {upload.file_type.upper()}")
    ws['A2'].font = Font(color='94A3B8')
    ws.merge_cells('A1:H1')
    ws.merge_cells('A2:H2')

    headers = ['Column', 'Type', 'Nulls %', 'Unique', 'Min', 'Max', 'Mean', 'Median']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    for ri, col in enumerate(analysis.get('columns', []), 5):
        row_data = [
            col['name'],
            'numeric' if col.get('is_numeric') else 'text',
            col.get('null_pct', 0),
            col.get('unique_count', 0),
            col.get('min', ''),
            col.get('max', ''),
            round(col.get('mean', 0) or 0, 4) if col.get('is_numeric') else '',
            round(col.get('median', 0) or 0, 4) if col.get('is_numeric') else '',
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0:
                cell.fill = alt_fill
                cell.font = Font(color='E2E0F0')

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # ── Sheet 2: Data Preview ─────────────────────────────────────
    ws2 = wb.create_sheet('Preview')
    preview = analysis.get('preview', {})
    if preview.get('columns'):
        for ci, col_name in enumerate(preview['columns'], 1):
            cell = ws2.cell(row=1, column=ci, value=col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
        for ri, row in enumerate(preview.get('rows', []), 2):
            for ci, val in enumerate(row, 1):
                ws2.cell(row=ri, column=ci, value=val)

    # ── Sheet 3: Correlation ──────────────────────────────────────
    corr = analysis.get('correlation')
    if corr:
        ws3 = wb.create_sheet('Correlation')
        for i, name in enumerate(corr['columns'], 2):
            ws3.cell(row=1, column=i, value=name).font = hdr_font
            ws3.cell(row=i, column=1, value=name).font = hdr_font
        for r, row in enumerate(corr['matrix'], 2):
            for c, val in enumerate(row, 2):
                cell = ws3.cell(row=r, column=c, value=val)
                if val is not None:
                    if val >= 0.7:    cell.fill = PatternFill('solid', fgColor='4C1D95')
                    elif val >= 0.4:  cell.fill = PatternFill('solid', fgColor='2D1B69')
                    elif val <= -0.4: cell.fill = PatternFill('solid', fgColor='7F1D1D')

    # ── Sheet 4: AI Insights ──────────────────────────────────────
    if upload.ai_insights:
        ws4 = wb.create_sheet('AI Insights')
        ws4['A1'] = 'AI-Generated Insights'
        ws4['A1'].font = Font(bold=True, size=12, color='A78BFA')
        ws4['A3'] = upload.ai_insights
        ws4['A3'].alignment = Alignment(wrap_text=True)
        ws4.column_dimensions['A'].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = upload.original_name.rsplit('.', 1)[0]
    resp  = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}_analysis.xlsx"'
    return resp




def _append_exception_tables_pdf(story, tables, h2_s, body_s, Table, TableStyle, colors):
    if not tables:
        return
    story.append(Paragraph('Decision tables', h2_s))
    for table in tables[:4]:
        story.append(Paragraph(table.get('title','Decision table'), body_s))
        cols = table.get('columns') or ['Entity','Metric','Support']
        rows = table.get('rows') or []
        data = [cols] + [[r.get('entity',''), r.get('metric',''), r.get('support','')] for r in rows[:8]]
        tbl = Table(data, colWidths=[5.0*cm, 3.0*cm, 7.0*cm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f1446')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('GRID',       (0,0), (-1,-1), 0.25, colors.HexColor('#e2e8f0')),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8fafc'), colors.white]),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        story.append(tbl)
    if governance.get('methodology'):
        story.append(Spacer(1, 8))
        story.append(Paragraph('Methodology summary', h2_s))
        for item in governance.get('methodology')[:3]:
            story.append(Paragraph(f'• {item}', body_s))
    if governance.get('caveats'):
        story.append(Spacer(1, 8))
        story.append(Paragraph('Known caveats', h2_s))
        for item in governance.get('caveats')[:4]:
            story.append(Paragraph(f'• {item}', body_s))
        story.append(Spacer(1, 10))


def _append_metadata_pdf(story, upload, ctx, h2_s, body_s, Table, TableStyle, colors):
    profile = ctx.get('profile') or {}
    analysis = ctx.get('analysis') or {}
    branding = ctx.get('branding') or build_branding_payload(user=request.user, request=request)
    story.append(PageBreak())
    story.append(Paragraph('Appendix & metadata', h2_s))
    governance = ctx.get('governance') or {}
    rows = [
        ['Source file', upload.original_name],
        ['Rows', str(profile.get('row_count') or upload.row_count or '')],
        ['Columns', str(profile.get('column_count') or upload.column_count or '')],
        ['Analysis type', ((profile.get('analysis_classification') or {}).get('analysis_type') or analysis.get('analysis_type') or '').replace('_',' ').title()],
        ['Primary dimension', ((profile.get('analysis_classification') or {}).get('primary_dimension') or '')],
        ['Primary measure', ((profile.get('analysis_classification') or {}).get('primary_measure') or '')],
        ['Generated', (governance.get('summary_lines') or ['—'])[0].replace('Generated: ', '') if governance.get('summary_lines') else '—'],
    ]
    tbl = Table([['Field','Value']] + rows, colWidths=[4.5*cm, 10.5*cm])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f1446')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('GRID',       (0,0), (-1,-1), 0.25, colors.HexColor('#e2e8f0')),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8fafc'), colors.white]),
    ]))
    story.append(tbl)
    if governance.get('methodology'):
        story.append(Spacer(1, 8))
        story.append(Paragraph('Methodology summary', h2_s))
        for item in governance.get('methodology')[:3]:
            story.append(Paragraph(f'• {item}', body_s))
    if governance.get('caveats'):
        story.append(Spacer(1, 8))
        story.append(Paragraph('Known caveats', h2_s))
        for item in governance.get('caveats')[:4]:
            story.append(Paragraph(f'• {item}', body_s))

# ── PDF export (with charts) ──────────────────────────────────────────────────

@login_required
def export_pdf(request, pk):
    upload = get_object_or_404(FileUpload, pk=pk, user=request.user)
    ctx = build_export_context(upload, mode=(request.GET.get('mode') or 'executive'), request=request)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable, Image,
                                        PageBreak)
    except ImportError:
        return HttpResponse('reportlab required for PDF export.', status=500)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.6*cm, rightMargin=1.6*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('T', parent=styles['Title'], textColor=colors.HexColor(branding.get('primary_color') or '#7c3aed'), fontSize=20, spaceAfter=8)
    sub_s   = ParagraphStyle('S', parent=styles['Normal'], textColor=colors.HexColor('#64748b'), fontSize=9, spaceAfter=10)
    h2_s    = ParagraphStyle('H2', parent=styles['Heading2'], textColor=colors.HexColor(branding.get('accent_color') or '#3b82f6'), fontSize=13, spaceBefore=14, spaceAfter=6)
    body_s  = ParagraphStyle('B', parent=styles['Normal'], textColor=colors.HexColor('#1e293b'), fontSize=9, leading=13, spaceAfter=4)
    story = []

    profile = ctx.get('profile') or {}
    dashboard = ctx.get('dashboard') or {}
    analysis = ctx.get('analysis') or {}
    branding = ctx.get('branding') or build_branding_payload(user=request.user, request=request)
    logo_path = branding.get('logo_path') or ''
    if logo_path:
        try:
            story.append(Image(logo_path, width=3.0*cm, height=1.2*cm))
            story.append(Spacer(1, 4))
        except Exception:
            pass
    story.append(Paragraph(branding.get('report_title') or upload.original_name, title_s))
    story.append(Paragraph(upload.original_name, body_s))
    story.append(Paragraph(f"Board export • {upload.row_count:,} rows • {upload.column_count} columns", sub_s))
    story.append(HRFlowable(color=colors.HexColor(branding.get('primary_color') or '#7c3aed'), thickness=1.2, spaceAfter=12))
    hero = dashboard.get('hero') or {}
    if hero:
        story.append(Paragraph(hero.get('headline') or 'Executive summary', h2_s))
        if hero.get('subheadline'):
            story.append(Paragraph(hero.get('subheadline'), body_s))
    for card in (ctx.get('decision_cards') or [])[:4]:
        title = f"{card.get('emoji','')} <b>{card.get('label','Signal')}</b> — {card.get('title','')}"
        body = card.get('body') or ''
        action = card.get('action') or ''
        story.append(Paragraph(title, body_s))
        if body:
            story.append(Paragraph(body, body_s))
        if action:
            story.append(Paragraph(f"Action: {action}", body_s))
        story.append(Spacer(1, 4))

    charts = ctx.get('charts') or []
    sections = ctx.get('sections') or []
    if sections:
        story.append(PageBreak())
        story.append(Paragraph('Section summaries', h2_s))
        for section in sections[:5]:
            story.append(Paragraph(section.get('title') or 'Section', body_s))
            if section.get('intro'):
                story.append(Paragraph(section.get('intro'), body_s))
            for chart in (section.get('charts') or [])[:2]:
                png = _chart_to_png(chart, width_in=6.6, height_in=2.6, branding=request.branding if hasattr(request, 'branding') else None)
                if png:
                    story.append(Image(io.BytesIO(png), width=17*cm, height=6.6*cm))
                explanation = getattr(chart, 'explanation', None)
                if explanation and explanation.get('summary'):
                    story.append(Paragraph(explanation.get('summary'), body_s))
                story.append(Spacer(1, 8))

    _append_exception_tables_pdf(story, ctx.get('exception_tables') or [], h2_s, body_s, Table, TableStyle, colors)
    _append_metadata_pdf(story, upload, ctx, h2_s, body_s, Table, TableStyle, colors)
    doc.build(story)
    buf.seek(0)
    fname = upload.original_name.rsplit('.',1)[0]
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}_board_export.pdf"'
    return resp


# ── PPTX export ───────────────────────────────────────────────────────────────

@login_required
def export_pptx(request, pk):
    upload = get_object_or_404(FileUpload, pk=pk, user=request.user)
    try:
        from .pptx_export import build_pptx
        data  = build_pptx(upload, request=request)
        fname = upload.original_name.rsplit('.', 1)[0]
        resp  = HttpResponse(
            data,
            content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation',
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname}_analysis.pptx"'
        return resp
    except Exception as e:
        return HttpResponse(f'PowerPoint export failed: {e}', status=500)


# ── Background export queue + status ─────────────────────────────────────────

def _job_payload(job):
    dl_url = '/export/download/%s/' % job.id if job.status == 'done' else ''
    return {
        'job_id': str(job.id),
        'status': job.status,
        'url': dl_url,
        'error': job.error or '',
        'fmt': job.fmt,
        'theme': job.theme,
        'created_at': job.created_at.isoformat() if getattr(job, 'created_at', None) else '',
        'updated_at': job.updated_at.isoformat() if getattr(job, 'updated_at', None) else '',
        'original_name': getattr(job.upload, 'original_name', ''),
    }


@login_required
def export_queue(request, pk, fmt):
    """
    Generate export synchronously in request thread.
    Returns job_id immediately; the job is already done by the time
    the client polls /export/status/<job_id>/.
    For large files the browser shows "generating…" for a few seconds then downloads.
    """
    from django.http import JsonResponse
    from .models import ExportJob
    from apps.exports.tasks import run_export

    upload = get_object_or_404(FileUpload, pk=pk, user=request.user)
    theme  = request.GET.get('theme', 'dark')
    job    = ExportJob.objects.create(
        user=request.user, upload=upload, fmt=fmt, theme=theme)

    # Run synchronously — file ready before response is sent
    run_export(str(job.id))
    job.refresh_from_db()

    payload = _job_payload(job)
    return JsonResponse(payload)


@login_required
def export_status(request, job_id):
    """Poll export job status."""
    from django.http import JsonResponse
    from .models import ExportJob
    try:
        job = ExportJob.objects.get(pk=job_id, user=request.user)
    except ExportJob.DoesNotExist:
        return JsonResponse({'status': 'error', 'error': 'Job not found'}, status=404)
    payload = _job_payload(job)
    return JsonResponse(payload)


@login_required
def export_download(request, job_id):
    """Stream the generated export file as an attachment download."""
    import os, mimetypes
    from .models import ExportJob
    try:
        job = ExportJob.objects.get(pk=job_id, user=request.user)
    except ExportJob.DoesNotExist:
        from django.http import Http404
        raise Http404

    if job.status != 'done' or not job.result_url:
        return HttpResponse('Export not ready yet.', status=202)

    # Build absolute path from media URL
    rel  = job.result_url.replace(settings.MEDIA_URL, '', 1)
    path = os.path.join(settings.MEDIA_ROOT, rel)

    if not os.path.exists(path):
        return HttpResponse('Export file not found.', status=404)

    ext_map = {'pdf': 'application/pdf',
               'pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation'}
    content_type = ext_map.get(job.fmt, 'application/octet-stream')

    base_name = job.upload.original_name.rsplit('.', 1)[0]
    fname     = '%s_%s.%s' % (base_name, job.fmt, job.fmt)

    with open(path, 'rb') as f:
        data = f.read()

    resp = HttpResponse(data, content_type=content_type)
    resp['Content-Disposition'] = 'attachment; filename="%s"' % fname
    resp['Content-Length']      = str(len(data))
    return resp


@login_required
def export_history(request, pk):
    """Return recent export jobs for a specific upload and user."""
    from django.http import JsonResponse
    from .models import ExportJob

    upload = get_object_or_404(FileUpload, pk=pk, user=request.user)
    jobs = ExportJob.objects.filter(user=request.user, upload=upload).select_related('upload')[:12]
    return JsonResponse({
        'jobs': [_job_payload(job) for job in jobs],
    })


@login_required
def export_retry(request, job_id):
    """Retry a previous export job with the same format and theme."""
    from django.http import JsonResponse
    from .models import ExportJob
    from apps.exports.tasks import run_export

    try:
        prev = ExportJob.objects.select_related('upload').get(pk=job_id, user=request.user)
    except ExportJob.DoesNotExist:
        return JsonResponse({'status': 'error', 'error': 'Job not found'}, status=404)

    retry_job = ExportJob.objects.create(
        user=request.user,
        upload=prev.upload,
        fmt=prev.fmt,
        theme=prev.theme or 'dark',
    )
    run_export(str(retry_job.id))
    retry_job.refresh_from_db()
    return JsonResponse(_job_payload(retry_job))
